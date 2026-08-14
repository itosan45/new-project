#!/usr/bin/env python3
"""書類の文字から項目を取り出して、Excelの決めた列に入れる。1ファイル完結版。

必要なもの: openpyxl だけ (pip install openpyxl)
インターネット接続もAPIの鍵も不要。

使い方:
    python3 転記.py 読み取り結果.json

読み取り結果.json の形:
    [{"ファイル名": "領収書01.jpg", "全文": "領収書\\n2026年8月12日\\n..."}]

列の構成を変えたいときは、下の「設定」だけ書き換えてください。
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

# ============================================================
# 設定 — ここだけ書き換えれば列構成を変えられます
# ============================================================

出力ファイル = "抽出結果.xlsx"
シート名 = "抽出データ"

# 種類に指定できるもの:
#   ファイル名 / 日付 / 金額 / 電話番号 / 全文 / キーワード
# 「キーワード」は 手がかり に書いた言葉の右または次の行の値を拾います
項目 = [
    {"列": "A", "見出し": "ファイル名", "種類": "ファイル名"},
    {"列": "B", "見出し": "日付", "種類": "日付"},
    {"列": "C", "見出し": "金額", "種類": "金額"},
    {"列": "D", "見出し": "宛名", "種類": "キーワード",
     "手がかり": ["宛名", "お名前", "氏名", "様"]},
    {"列": "E", "見出し": "電話番号", "種類": "電話番号"},
    {"列": "F", "見出し": "読み取り全文", "種類": "全文"},
]

# ============================================================
# ここから下は書き換え不要
# ============================================================

_日付パターン = [
    re.compile(r"(\d{4})\s*[/\-年]\s*(\d{1,2})\s*[/\-月]\s*(\d{1,2})\s*日?"),
    re.compile(r"令和\s*(\d{1,2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"),
]
_金額パターン = re.compile(r"[¥￥]\s*([0-9][0-9,]*)|([0-9][0-9,]*)\s*円")
_電話パターン = re.compile(r"0\d{1,4}[-(）\s]?\d{1,4}[-)）\s]?\d{3,4}")
_敬称 = ("様", "さま", "サマ", "さん", "御中", "殿")


def 日付を探す(文字列):
    for i, パターン in enumerate(_日付パターン):
        m = パターン.search(文字列)
        if not m:
            continue
        年, 月, 日 = (int(x) for x in m.groups())
        if i == 1:
            年 += 2018  # 令和 → 西暦
        if not (1 <= 月 <= 12 and 1 <= 日 <= 31):
            continue
        return f"{年:04d}-{月:02d}-{日:02d}"
    return ""


def 金額を探す(文字列):
    """複数あれば一番大きい額を返す。合計であることが多いため。"""
    金額一覧 = []
    for m in _金額パターン.finditer(文字列):
        数字 = m.group(1) or m.group(2)
        try:
            金額一覧.append(int(数字.replace(",", "")))
        except ValueError:
            continue
    return str(max(金額一覧)) if 金額一覧 else ""


def 電話番号を探す(文字列):
    m = _電話パターン.search(文字列)
    return m.group(0).strip() if m else ""


def _敬称を落とす(値):
    for 敬 in _敬称:
        if 値.endswith(敬):
            return 値[: -len(敬)].strip()
    return 値


def キーワードで探す(文字列, 手がかり一覧):
    行一覧 = [行.strip() for 行 in 文字列.splitlines()]
    for i, 行 in enumerate(行一覧):
        for 手がかり in 手がかり一覧:
            if 手がかり not in 行:
                continue
            右側 = 行.split(手がかり, 1)[1].strip(" :：　|｜")
            if 右側:
                return _敬称を落とす(右側)
            if i + 1 < len(行一覧) and 行一覧[i + 1]:
                return _敬称を落とす(行一覧[i + 1])
    return ""


def 一件ぶんを抽出する(全文, ファイル名):
    結果 = {}
    for 設定 in 項目:
        種類 = 設定["種類"]
        if 種類 == "ファイル名":
            値 = ファイル名
        elif 種類 == "日付":
            値 = 日付を探す(全文)
        elif 種類 == "金額":
            値 = 金額を探す(全文)
        elif 種類 == "電話番号":
            値 = 電話番号を探す(全文)
        elif 種類 == "全文":
            値 = 全文
        elif 種類 == "キーワード":
            値 = キーワードで探す(全文, 設定.get("手がかり", []))
        else:
            値 = ""
        結果[設定["列"]] = 値
    return 結果


def 次の空き行を探す(シート):
    行 = シート.max_row
    while 行 > 0 and all(
        シート.cell(row=行, column=列).value in (None, "")
        for 列 in range(1, シート.max_column + 1)
    ):
        行 -= 1
    return 行 + 1


def main(引数):
    if len(引数) != 1:
        print("使い方: python3 転記.py 読み取り結果.json")
        return 1

    入力パス = Path(引数[0])
    if not 入力パス.exists():
        print(f"ファイルが見つかりません: {入力パス}")
        return 1

    with 入力パス.open(encoding="utf-8") as f:
        読み取り結果 = json.load(f)
    if not 読み取り結果:
        print("読み取り結果が空です。")
        return 0

    出力パス = Path(出力ファイル)
    if 出力パス.exists():
        ブック = load_workbook(出力パス)
        シート = (ブック[シート名] if シート名 in ブック.sheetnames
                  else ブック.create_sheet(シート名))
    else:
        ブック = Workbook()
        シート = ブック.active
        シート.title = シート名

    if シート.max_row <= 1 and not シート["A1"].value:
        for 設定 in 項目:
            シート[f"{設定['列']}1"] = 設定["見出し"]

    for 一件 in 読み取り結果:
        値 = 一件ぶんを抽出する(一件.get("全文", ""), 一件.get("ファイル名", ""))
        行 = 次の空き行を探す(シート)
        for 列, 中身 in 値.items():
            シート[f"{列}{行}"] = 中身
        print(f"  {行}行目 ← {一件.get('ファイル名', '')}")

    ブック.save(出力パス)
    print(f"\n{len(読み取り結果)}件を書き込みました: {出力パス.resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
