"""抽出した値をExcelに書き込む処理。"""

from pathlib import Path

from openpyxl import Workbook, load_workbook


def ブックを開くか作る(パス: Path, シート名: str):
    """既存のExcelがあれば開き、無ければ新しく作る。"""
    if パス.exists():
        ブック = load_workbook(パス)
        シート = ブック[シート名] if シート名 in ブック.sheetnames else ブック.create_sheet(シート名)
    else:
        ブック = Workbook()
        シート = ブック.active
        シート.title = シート名
    return ブック, シート


def 見出しを書く(シート, 項目設定: list[dict]) -> None:
    """1行目に見出しを書く。すでに何か書かれていれば触らない。"""
    if シート.max_row > 1 or シート["A1"].value:
        return
    for 項目 in 項目設定:
        シート[f"{項目['列']}1"] = 項目["見出し"]


def 次の空き行を探す(シート) -> int:
    """一番下の、まだ何も書かれていない行の番号を返す。"""
    行 = シート.max_row
    while 行 > 0 and all(
        シート.cell(row=行, column=列).value in (None, "")
        for 列 in range(1, シート.max_column + 1)
    ):
        行 -= 1
    return 行 + 1


def 一行書き込む(シート, 行番号: int, 値: dict) -> None:
    """{列: 値} を指定した行に書き込む。"""
    for 列, 中身 in 値.items():
        シート[f"{列}{行番号}"] = 中身
