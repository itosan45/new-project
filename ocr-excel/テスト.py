#!/usr/bin/env python3
"""抽出とExcel書き込みの動作確認。Vision APIは使わないので無料で実行できる。

    python3 テスト.py
"""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from lib import エクセル, 抽出  # noqa: E402


class 抽出のテスト(unittest.TestCase):
    def test_西暦の日付を取り出せる(self):
        self.assertEqual(抽出.日付を探す("2026年8月12日 領収書"), "2026-08-12")
        self.assertEqual(抽出.日付を探す("発行日 2026/8/5"), "2026-08-05")
        self.assertEqual(抽出.日付を探す("2026-12-31"), "2026-12-31")

    def test_令和の日付を西暦に直せる(self):
        self.assertEqual(抽出.日付を探す("令和8年8月12日"), "2026-08-12")

    def test_日付がなければ空(self):
        self.assertEqual(抽出.日付を探す("ただの文章です"), "")

    def test_ありえない月日は拾わない(self):
        self.assertEqual(抽出.日付を探す("2026年13月45日"), "")

    def test_金額を取り出せる(self):
        self.assertEqual(抽出.金額を探す("合計 ¥12,300"), "12300")
        self.assertEqual(抽出.金額を探す("1,500円"), "1500")

    def test_複数の金額なら一番大きいものを返す(self):
        文 = "小計 1,000円\n消費税 100円\n合計 1,100円"
        self.assertEqual(抽出.金額を探す(文), "1100")

    def test_電話番号を取り出せる(self):
        self.assertEqual(抽出.電話番号を探す("TEL 03-1234-5678"), "03-1234-5678")
        self.assertEqual(抽出.電話番号を探す("090-1111-2222まで"), "090-1111-2222")

    def test_キーワードの右側を拾える(self):
        self.assertEqual(
            抽出.キーワードで探す("宛名: 山田太郎", ["宛名"]), "山田太郎"
        )

    def test_宛名の敬称を落とす(self):
        self.assertEqual(抽出.キーワードで探す("宛名: 山田太郎 様", ["宛名"]), "山田太郎")
        self.assertEqual(抽出.キーワードで探す("宛名\n株式会社ABC 御中", ["宛名"]), "株式会社ABC")

    def test_キーワードの次の行を拾える(self):
        self.assertEqual(
            抽出.キーワードで探す("お名前\n鈴木花子\n住所", ["お名前"]), "鈴木花子"
        )

    def test_設定どおりに一件分を組み立てられる(self):
        全文 = "領収書\n2026年8月12日\n宛名: 山田太郎\n合計 ¥3,000\nTEL 03-1111-2222"
        項目設定 = [
            {"列": "A", "見出し": "ファイル名", "種類": "ファイル名"},
            {"列": "B", "見出し": "日付", "種類": "日付"},
            {"列": "C", "見出し": "金額", "種類": "金額"},
            {"列": "D", "見出し": "宛名", "種類": "キーワード", "手がかり": ["宛名"]},
            {"列": "E", "見出し": "電話", "種類": "電話番号"},
        ]
        結果 = 抽出.一件ぶんを抽出する(全文, "領収書01.jpg", 項目設定)
        self.assertEqual(結果["A"], "領収書01.jpg")
        self.assertEqual(結果["B"], "2026-08-12")
        self.assertEqual(結果["C"], "3000")
        self.assertEqual(結果["D"], "山田太郎")
        self.assertEqual(結果["E"], "03-1111-2222")


class エクセルのテスト(unittest.TestCase):
    def test_新規作成して書き込める(self):
        項目設定 = [
            {"列": "A", "見出し": "ファイル名", "種類": "ファイル名"},
            {"列": "B", "見出し": "金額", "種類": "金額"},
        ]
        with tempfile.TemporaryDirectory() as 一時:
            パス = Path(一時) / "テスト.xlsx"
            ブック, シート = エクセル.ブックを開くか作る(パス, "データ")
            エクセル.見出しを書く(シート, 項目設定)
            エクセル.一行書き込む(シート, 2, {"A": "a.jpg", "B": "100"})
            ブック.save(パス)

            self.assertTrue(パス.exists())
            from openpyxl import load_workbook

            確認 = load_workbook(パス)["データ"]
            self.assertEqual(確認["A1"].value, "ファイル名")
            self.assertEqual(確認["A2"].value, "a.jpg")

    def test_追記すると次の行に入る(self):
        項目設定 = [{"列": "A", "見出し": "ファイル名", "種類": "ファイル名"}]
        with tempfile.TemporaryDirectory() as 一時:
            パス = Path(一時) / "テスト.xlsx"

            ブック, シート = エクセル.ブックを開くか作る(パス, "データ")
            エクセル.見出しを書く(シート, 項目設定)
            エクセル.一行書き込む(シート, エクセル.次の空き行を探す(シート), {"A": "1枚目"})
            ブック.save(パス)

            ブック, シート = エクセル.ブックを開くか作る(パス, "データ")
            行 = エクセル.次の空き行を探す(シート)
            エクセル.一行書き込む(シート, 行, {"A": "2枚目"})
            ブック.save(パス)

            from openpyxl import load_workbook

            確認 = load_workbook(パス)["データ"]
            self.assertEqual(確認["A2"].value, "1枚目")
            self.assertEqual(確認["A3"].value, "2枚目")


if __name__ == "__main__":
    unittest.main(verbosity=2)
